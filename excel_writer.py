from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers


HEADERS = [
    "日期", "股票代码", "股票名称", "交易所", "数量",
    "单价", "BUY/SELL", "费用", "货币", "总净额",
]

COL_WIDTHS = [12, 12, 18, 12, 10, 12, 10, 12, 8, 14]


def create_excel(rows: List[dict], output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "交易明细"

    # Header row
    header_font = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data rows
    number_fmt = "#,##0.00"
    int_fmt = "#,##0"

    for row_idx, row in enumerate(rows, 2):
        date_cell = ws.cell(row=row_idx, column=1, value=datetime.strptime(row["date"], "%Y/%m/%d"))
        date_cell.number_format = "YYYY/MM/DD"
        code = row["stock_code"]
        code_cell = ws.cell(row=row_idx, column=2, value=int(code) if code.isdigit() else code)
        if code.isdigit():
            code_cell.number_format = "0"
        ws.cell(row=row_idx, column=3, value=row["stock_name"])
        ws.cell(row=row_idx, column=4, value=row["exchange"])

        qty_cell = ws.cell(row=row_idx, column=5, value=row["quantity"])
        qty_cell.number_format = int_fmt

        price_cell = ws.cell(row=row_idx, column=6, value=row["unit_price"])
        price_cell.number_format = number_fmt

        ws.cell(row=row_idx, column=7, value=row["direction"])

        fee_cell = ws.cell(row=row_idx, column=8, value=row["fees"])
        fee_cell.number_format = number_fmt

        ws.cell(row=row_idx, column=9, value=row["currency"])

        net_cell = ws.cell(row=row_idx, column=10, value=row["total_net_amount"])
        net_cell.number_format = number_fmt

    wb.save(output_path)
    return output_path
